import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import date, datetime
import time

# --- 1. 系統初始化 ---
st.set_page_config(page_title="全店業績戰情室", layout="wide", page_icon="📈")

# 初始化 Session State
if 'preview_data' not in st.session_state: st.session_state.preview_data = None
if 'preview_score' not in st.session_state: st.session_state.preview_score = 0
if 'authenticated_store' not in st.session_state: st.session_state.authenticated_store = None
if 'current_excel_file' not in st.session_state: st.session_state.current_excel_file = None
if 'admin_logged_in' not in st.session_state: st.session_state.admin_logged_in = False

# 檢查必要設定
if "gcp_service_account" not in st.secrets:
    st.error("❌ 嚴重錯誤：Secrets 中找不到 [gcp_service_account]。")
    st.stop()
if "TARGET_FOLDER_ID" not in st.secrets:
    st.warning("⚠️ 警告：Secrets 中找不到 TARGET_FOLDER_ID。")

# Google 套件
try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
except ImportError:
    st.error("❌ 缺少 Google 套件，請檢查 requirements.txt")
    st.stop()

# --- 2. Google Drive 功能 (核心) ---
def get_drive_service():
    creds_dict = dict(st.secrets["gcp_service_account"])
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=['https://www.googleapis.com/auth/drive']
    )
    return build('drive', 'v3', credentials=creds)

def get_file_id_in_folder(service, filename, folder_id):
    """全域搜尋檔案"""
    query = f"name = '{filename}' and trashed = false"
    # 增加 parents 查詢以確保在正確資料夾
    if folder_id:
        query += f" and '{folder_id}' in parents"
        
    results = service.files().list(q=query, fields="files(id, name)", orderBy="createdTime desc").execute()
    items = results.get('files', [])
    if not items: return None
    return items[0]['id']

def update_excel_drive(store, staff, date_obj, data_dict):
    """寫入資料到雲端 Excel"""
    folder_id = st.secrets.get("TARGET_FOLDER_ID")
    filename = f"{date_obj.year}_{date_obj.month:02d}_{store}業績日報表.xlsx"
    
    try:
        service = get_drive_service()
        file_id = get_file_id_in_folder(service, filename, folder_id)
        if not file_id:
            return f"❌ 找不到檔案 [{filename}]，請確認雲端硬碟檔名。"

        request = service.files().get_media(fileId=file_id)
        file_content = request.execute()
        excel_stream = BytesIO(file_content)
        
        wb = openpyxl.load_workbook(excel_stream)
        if staff not in wb.sheetnames:
            return f"❌ 找不到人員分頁：[{staff}]"
        
        ws = wb[staff]
        # 假設第 15 列是 1 號，則當日列數為 15 + (日期 - 1)
        target_row = 15 + (date_obj.day - 1)
        
        # 定義欄位對應 (依據 Excel 實際欄位順序調整)
        col_map = {
            '毛利': 2, 
            '門號': 3, 
            '保險營收': 4, 
            '配件營收': 5,
            '庫存手機': 6, 
            '蘋果手機': 7, 
            '蘋果平板+手錶': 8, 
            'VIVO手機': 9,
            '生活圈': 10, 
            'GOOGLE 評論': 11, 
            '來客數': 12,
            '遠傳續約': 13,        # 新增
            '遠傳續約累積GAP': 14, # 順延
            '遠傳升續率': 15,      # 順延
            '遠傳平續率': 16,      # 順延
            '綜合指標': 17         # 新增
        }
        
        # 這些欄位是直接覆蓋數值 (不是累加)
        overwrite_fields = ['遠傳續約累積GAP', '遠傳升續率', '遠傳平續率', '綜合指標']
        
        for field, new_val in data_dict.items():
            if field in col_map and new_val is not None:
                col_idx = col_map[field]
                cell = ws.cell(row=target_row, column=col_idx)
                
                # 讀取舊值 (若非數值則設為 0)
                old_val = cell.value if isinstance(cell.value, (int, float)) else 0
                
                if field in overwrite_fields:
                    cell.value = new_val
                else:
                    # 其他欄位採累加模式 (可依需求改為覆蓋)
                    cell.value = old_val + new_val

        output_stream = BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        media = MediaIoBaseUpload(output_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        service.files().update(fileId=file_id, media_body=media).execute()
        
        return f"✅ 資料已成功寫入：{filename}"

    except Exception as e:
        return f"❌ 系統錯誤: {str(e)}"

def read_excel_drive(store, date_obj):
    """回傳：(檔案內容Bytes, 檔名, 線上連結URL)"""
    folder_id = st.secrets.get("TARGET_FOLDER_ID")
    filename = f"{date_obj.year}_{date_obj.month:02d}_{store}業績日報表.xlsx"
    
    try:
        service = get_drive_service()
        file_id = get_file_id_in_folder(service, filename, folder_id)
        
        if not file_id:
            return None, f"找不到檔案：{filename}", None

        # 1. 取得檔案的線上連結 (webViewLink)
        file_meta = service.files().get(fileId=file_id, fields='webViewLink').execute()
        file_url = file_meta.get('webViewLink')

        # 2. 下載檔案內容
        request = service.files().get_media(fileId=file_id)
        file_content = request.execute()
        
        return file_content, filename, file_url

    except Exception as e:
        return None, str(e), None

def aggregate_all_stores(date_obj):
    """(新增功能) 彙整所有分店當月數據"""
    folder_id = st.secrets.get("TARGET_FOLDER_ID")
    service = get_drive_service()
    
    all_data = []
    
    # 遍歷所有分店
    for store_name in STORES.keys():
        if store_name == "(ALL) 全店總表": continue
        
        filename = f"{date_obj.year}_{date_obj.month:02d}_{store_name}業績日報表.xlsx"
        file_id = get_file_id_in_folder(service, filename, folder_id)
        
        store_stats = {
            "門市": store_name,
            "毛利": 0, "門號": 0, "保險營收": 0, "配件營收": 0,
            "來客數": 0, "遠傳續約": 0, "綜合指標": 0,
            "連結": None
        }

        if file_id:
            # 取得連結
            meta = service.files().get(fileId=file_id, fields='webViewLink').execute()
            store_stats["連結"] = meta.get('webViewLink')
            
            # 讀取內容進行簡單加總 (這裡只示範讀取 '總表' 分頁的最後一列，或是累加所有人員)
            # 為了效能，這裡暫時只讀取檔案存在與否，若要深入讀取數值需下載每個 Excel
            # 這裡示範：標記為「已讀取」
            store_stats["狀態"] = "✅ 線上"
        else:
            store_stats["狀態"] = "❌ 未建立"
            
        all_data.append(store_stats)
        
    return pd.DataFrame(all_data)

# --- 3. 組織與目標 ---
STORES = {
    "(ALL) 全店總表": [],
    "文賢店": ["慧婷", "阿緯", "子翔", "默默"],
    "東門店": ["小萬", "914", "默默", "人員4"],
    "永康店": ["宗憲", "筑君", "澤偉", "翰霖", "77", "支援"],
    "歸仁店": ["配飯", "誌廷", "阿孝", "支援", "人員2"],
    "安中店": ["宗憲", "大俗", "翰霖", "澤偉"],
    "小西門店": ["豆豆", "秀秀", "人員3", "人員4"],
    "鹽行店": ["配飯", "薪融", "脆迪", "誌廷", "人員2"],
    "五甲店": ["阿凱", "孟婧", "支援", "人員2"],
    "鳳山店": ["店長", "組員"]
}
# 預設目標 (可依需求調整)
DEFAULT_TARGETS = {'毛利': 140000, '門號': 24, '保險': 28000, '配件': 35000, '庫存': 21}

# --- 4. 介面與權限邏輯 ---

st.sidebar.title("🏢 門市導航")
selected_store = st.sidebar.selectbox("請選擇門市", list(STORES.keys()))

if selected_store == "(ALL) 全店總表":
    staff_options = []
    selected_user = "全店總覽"
else:
    staff_options = ["該店總表"] + STORES[selected_store]
    selected_user = st.sidebar.selectbox("請選擇人員", staff_options)

st.title(f"📊 {selected_store} - {selected_user}")

# 權限驗證函式
def check_store_auth(current_store):
    # 全店總表 -> 管理員密碼
    if current_store == "(ALL) 全店總表":
        if st.session_state.admin_logged_in: return True
        st.info("🛡️ 此區域需要管理員權限")
        admin_input = st.text_input("🔑 請輸入管理員密碼", type="password", key="admin_input")
        if st.button("驗證管理員"):
            if admin_input == st.secrets.get("admin_password"):
                st.session_state.admin_logged_in = True
                st.rerun()
            else:
                st.error("❌ 密碼錯誤")
        return False

    # 各分店 -> 分店密碼
    if st.session_state.authenticated_store == current_store: return True

    st.info(f"🔒 請輸入【{current_store}】的專屬密碼")
    with st.form("store_login"):
        input_pass = st.text_input("密碼", type="password")
        login_btn = st.form_submit_button("登入")
        if login_btn:
            # 從 secrets["store_passwords"] 取得密碼
            correct_pass = st.secrets["store_passwords"].get(current_store)
            if not correct_pass: st.error("⚠️ 未設定密碼")
            elif input_pass == correct_pass:
                st.session_state.authenticated_store = current_store
                st.success("登入成功！")
                st.rerun()
            else: st.error("❌ 密碼錯誤")
    return False

if not check_store_auth(selected_store):
    st.stop()

# =========================================================
# 主畫面邏輯
# =========================================================

if selected_store == "(ALL) 全店總表":
    st.success("✅ 管理員權限已解鎖")
    st.markdown("### 🏆 全公司業績戰情室")
    
    col_date, _ = st.columns([1, 3])
    view_date = col_date.date_input("選擇檢視月份", date.today())
    
    if st.button("🔄 讀取全部分店數據"):
        with st.spinner("正在連線各分店報表..."):
            df_all_stores = aggregate_all_stores(view_date)
            st.dataframe(
                df_all_stores, 
                column_config={
                    "連結": st.column_config.LinkColumn("雲端檔案")
                },
                use_container_width=True
            )
            st.caption("💡 提示：點擊連結可直接開啟各店原始 Excel 檔")

elif selected_user == "該店總表":
    # ----------------------------------------------------
    # 門市報表檢視中心 (含線上連結)
    # ----------------------------------------------------
    st.markdown("### 📥 門市報表檢視中心")
    st.info("在此您可以下載、線上預覽，或直接開啟 Google 試算表。")

    col_d1, col_d2 = st.columns([1, 2])
    view_date = col_d1.date_input("選擇報表月份", date.today())
    
    if col_d1.button("📂 讀取雲端報表", use_container_width=True):
        with st.spinner("正在從 Google Drive 讀取資料..."):
            file_bytes, file_msg, file_link = read_excel_drive(selected_store, view_date)
            
            if file_bytes:
                st.session_state.current_excel_file = {
                    'bytes': file_bytes,
                    'name': file_msg,
                    'link': file_link
                }
                st.success("✅ 報表讀取成功！")
            else:
                st.error(f"❌ {file_msg}")
    
    # 顯示操作區
    if st.session_state.current_excel_file:
        file_data = st.session_state.current_excel_file
        st.divider()
        st.subheader(f"📄 檔案：{file_data['name']}")
        
        # 三大按鈕
        c_btn1, c_btn2, c_btn3 = st.columns(3)
        
        # 1. Google Drive 開啟連結
        if file_data.get('link'):
            c_btn1.link_button(
                "🔗 在 Google Drive 開啟", 
                file_data['link'], 
                type="primary", 
                use_container_width=True
            )
        
        # 2. 下載按鈕
        c_btn2.download_button(
            label="💾 下載 Excel 檔",
            data=file_data['bytes'],
            file_name=file_data['name'],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        # 3. 重整
        if c_btn3.button("🔄 重新整理數據", use_container_width=True):
            st.session_state.current_excel_file = None
            st.rerun()

        st.markdown("---")
        st.write("#### 👀 網頁內快速預覽 (唯讀)")
        
        try:
            excel_obj = pd.ExcelFile(BytesIO(file_data['bytes']))
            sheet_names = excel_obj.sheet_names
            col_sheet, _ = st.columns([1, 2])
            selected_sheet = col_sheet.selectbox("選擇要檢視的分頁", sheet_names)
            
            df_preview = pd.read_excel(excel_obj, sheet_name=selected_sheet)
            st.dataframe(df_preview, use_container_width=True)
            
        except Exception as e:
            st.warning("預覽載入失敗，請直接開啟檔案查看。")

else:
    # ----------------------------------------------------
    # 個人填寫模式 (Step 1 預覽 -> Step 2 上傳)
    # ----------------------------------------------------
    st.markdown("### 📝 今日業績回報")

    with st.form("daily_input_full"):
        d_col1, d_col2 = st.columns([1, 3])
        input_date = d_col1.date_input("📅 報表日期", date.today())
        st.markdown("---")

        # 1. 財務與門號
        st.subheader("💰 財務與門號 (Core)")
        c1, c2, c3, c4 = st.columns(4)
        in_profit = c1.number_input("毛利 ($)", min_value=0, step=100)
        in_number = c2.number_input("門號 (件)", min_value=0, step=1)
        in_insur = c3.number_input("保險營收 ($)", min_value=0, step=100)
        in_acc = c4.number_input("配件營收 ($)", min_value=0, step=100)

        # 2. 硬體銷售
        st.subheader("📱 硬體銷售 (Hardware)")
        h1, h2, h3, h4 = st.columns(4)
        in_stock = h1.number_input("庫存手機 (台)", min_value=0, step=1)
        in_vivo = h2.number_input("VIVO 手機 (台)", min_value=0, step=1)
        in_apple = h3.number_input("🍎 蘋果手機 (台)", min_value=0, step=1)
        in_ipad = h4.number_input("🍎 平板/手錶 (台)", min_value=0, step=1)

        # 3. 顧客經營
        st.subheader("🤝 顧客經營 (Service)")
        s1, s2, s3 = st.columns(3)
        in_life = s1.number_input("生活圈 (件)", min_value=0, step=1)
        in_review = s2.number_input("Google 評論 (則)", min_value=0, step=1)
        in_traffic = s3.number_input("來客數 (人)", min_value=0, step=1)

        # 4. 遠傳專案指標
        st.subheader("📡 遠傳專案指標 (KPI)")
        t1, t2, t3, t4 = st.columns(4)
        in_renew = t1.number_input("遠傳續約 (件)", min_value=0, step=1)
        in_gap = t2.number_input("遠傳續約累積 GAP", step=1)
        in_up_rate_raw = t3.number_input("遠傳升續率 (%)", min_value=0.0, max_value=100.0, step=0.1)
        in_flat_rate_raw = t4.number_input("遠傳平續率 (%)", min_value=0.0, max_value=100.0, step=0.1)
        
        # 5. 綜合指標
        st.subheader("🏆 綜合評估")
        in_composite = st.number_input("綜合指標分數", min_value=0.0, step=0.1)
        
        check_btn = st.form_submit_button("🔍 試算分數並預覽 (Step 1)", use_container_width=True)

        if check_btn:
            # 簡易試算邏輯 (可自訂)
            def calc(act, tgt, w): return (act / tgt * w) if tgt > 0 else 0
            score = (
                calc(in_profit, DEFAULT_TARGETS['毛利'], 0.25) + 
                calc(in_number, DEFAULT_TARGETS['門號'], 0.20) + 
                calc(in_insur, DEFAULT_TARGETS['保險'], 0.15) + 
                calc(in_acc, DEFAULT_TARGETS['配件'], 0.15) + 
                calc(in_stock, DEFAULT_TARGETS['庫存'], 0.15)
            )
            
            st.session_state.preview_data = {
                '毛利': in_profit, '門號': in_number, '保險營收': in_insur, '配件營收': in_acc,
                '庫存手機': in_stock, '蘋果手機': in_apple, '蘋果平板+手錶': in_ipad, 'VIVO手機': in_vivo,
                '生活圈': in_life, 'GOOGLE 評論': in_review, '來客數': in_traffic,
                '遠傳續約': in_renew,
                '遠傳續約累積GAP': in_gap, 
                '遠傳升續率': in_up_rate_raw / 100, 
                '遠傳平續率': in_flat_rate_raw / 100,
                '綜合指標': in_composite,
                '日期': input_date
            }
            st.session_state.preview_score = score
            st.rerun()

    if st.session_state.preview_data:
        st.divider()
        st.markdown("### 👀 請確認下方資料是否正確？")
        
        # 顯示預覽表格，並格式化百分比
        df_preview = pd.DataFrame([st.session_state.preview_data])
        # 隱藏日期欄位以免混淆
        display_df = df_preview.drop(columns=['日期'])
        
        st.dataframe(
            display_df, 
            hide_index=True,
            column_config={
                "遠傳升續率": st.column_config.NumberColumn(format="%.1f%%"),
                "遠傳平續率": st.column_config.NumberColumn(format="%.1f%%"),
                "毛利": st.column_config.NumberColumn(format="$%d"),
            }
        )
        
        if st.session_state.preview_score > 0:
            st.info(f"💡 系統試算核心貢獻度：{st.session_state.preview_score*100:.1f} 分 (僅供參考)")
        
        col_confirm, col_cancel = st.columns([1, 1])
        if col_confirm.button("✅ 確認無誤，立即上傳 (Step 2)", type="primary", use_container_width=True):
            progress_text = "連線 Google Drive 中...請稍候"
            my_bar = st.progress(0, text=progress_text)
            try:
                data_to_save = st.session_state.preview_data.copy()
                target_date = data_to_save.pop('日期')
                my_bar.progress(30, text="正在搜尋雲端檔案...")
                
                result_msg = update_excel_drive(selected_store, selected_user, target_date, data_to_save)
                my_bar.progress(100, text="處理完成！")
                
                if "✅" in result_msg:
                    st.success(result_msg)
                    st.balloons()
                    st.session_state.preview_data = None
                    st.session_state.preview_score = 0
                    time.sleep(3)
                    st.rerun()
                else:
                    st.error(result_msg)
            except Exception as e:
                st.error(f"❌ 錯誤: {str(e)}")
        
        if col_cancel.button("❌ 有錯誤，重新填寫", use_container_width=True):
            st.session_state.preview_data = None
            st.rerun()
